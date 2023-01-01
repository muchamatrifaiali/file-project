#Sure, here is an example of the code you need for Python to generate XML file from Excel:

import openpyxl 
from openpyxl import load_workbook 
from yattag import Doc, indent 
# load the excel file along with path where exists 
wb = load_workbook("empdata.xlsx") 
ws = wb.worksheets[0] 
# Create Yattag doc, tag and text objects 
doc, tag, text = Doc().tagtext() 
xml_header = '' 
xml_schema = '' 
doc.asis(xml_header) 
doc.asis(xml_schema) 
with tag('Employees'): 
# Use ws.max_row for all rows 
for row in ws.iter_rows(min_row=2, max_row=100, min_col=1, max_col=12): 
row = [cell.value for cell in row] 
with tag("Employee"): 
with tag("FirstName"): 
text(row[0]) 
with tag("LastName"): 
text(row[1]) 
with tag("Email"): 
text(row[10]) 
with tag("Phone"): 
text(row[8]) 
with tag("Company"): 
text(row[2]) 
result = indent( doc.getvalue(), indentation = ' ', indent_text = True) 
print(result) 
with open("employee.xml", "w") as f: #give path where you want to create 
f.write(result).


